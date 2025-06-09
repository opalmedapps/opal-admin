# SPDX-FileCopyrightText: Copyright (C) 2024 Opal Health Informatics Group at the Research Institute of the McGill University Health Centre <john.kildea@mcgill.ca>
#
# SPDX-License-Identifier: AGPL-3.0-or-later

import csv
import datetime as dt
import io
import zipfile
from http import HTTPStatus
from typing import TYPE_CHECKING, Any, cast

from django.http import FileResponse
from django.http.response import HttpResponseBadRequest
from django.test import Client
from django.urls.base import reverse
from django.utils import timezone
from django.utils.text import Truncator

import pytest
from openpyxl import load_workbook
from pytest_django.asserts import assertContains, assertTemplateUsed
from pytest_mock import MockerFixture

from opal.caregivers import factories as caregiver_factories
from opal.caregivers import models as caregiver_models
from opal.core import utils
from opal.legacy import factories as legacy_factories
from opal.patients import factories as patient_factories
from opal.patients import models as patient_models
from opal.usage_statistics import factories as stats_factories
from opal.users import factories as user_factories
from opal.users.models import Caregiver, User

from ..common import GroupByComponent, GroupReportType
from ..forms import GroupUsageStatisticsForm, IndividualUsageStatisticsForm

if TYPE_CHECKING:
    from collections.abc import Iterator

pytestmark = pytest.mark.django_db(databases=['default', 'legacy'])


# Content of empty summary reports
empty_summary_report_expected_file_contents = {
    'registration_summary': [
        ['uncompleted_registration', 'completed_registration', 'total_registration_codes'],
        ['0', '0', '0'],
    ],
    'grouped_registration_summary': [],
    'caregivers_summary': [
        [
            'caregivers_total',
            'caregivers_registered',
            'caregivers_unregistered',
            'never_logged_in_after_registration',
            'en',
            'fr',
        ],
        ['0', '0', '0', '0', '0', '0'],
    ],
    'fetch_patients_summary': [
        ['total', 'deceased', 'male', 'female', 'sex_other', 'sex_unknown', 'access_all', 'access_ntk'],
        ['0', '0', '0', '0', '0', '0', '0', '0'],
    ],
    'devices_summary': [
        ['device_total', 'device_ios', 'device_android', 'device_browser'],
        ['0', '0', '0', '0'],
    ],
}

# Content of empty received data reports
empty_received_data_report_expected_file_contents = {
    'received_questionnaires_summary': [],
    'received_documents_summary': [],
    'received_educational_materials_summary': [],
    'received_appointments_summary': [],
    'received_labs_summary': [],
    'received_clinical_data_summary': [
        [
            'no_appointments_labs_notes',
            'has_appointments_only',
            'has_labs_only',
            'has_clinical_notes_only',
            'receiving_new_data_total',
        ],
        ['0', '0', '0', '0', '0'],
    ],
}

# Content of empty app activity data reports
empty_app_activity_report_expected_file_contents: dict[str, Any] = {
    'users_latest_login_year_summary': [],
    'user_patient_clicks_summary': [],
    'users_clicks_summary': [],
    'logins_summary': [],
}

empty_group_report_filters_data = [
    (
        {
            'group_by': GroupByComponent.DAY.name,
            'report_type': GroupReportType.SUMMARY_REPORT.name,
        },
        empty_summary_report_expected_file_contents,
    ),
    (
        {
            'group_by': GroupByComponent.MONTH.name,
            'report_type': GroupReportType.SUMMARY_REPORT.name,
        },
        empty_summary_report_expected_file_contents,
    ),
    (
        {
            'group_by': GroupByComponent.YEAR.name,
            'report_type': GroupReportType.SUMMARY_REPORT.name,
        },
        empty_summary_report_expected_file_contents,
    ),
    (
        {
            'group_by': GroupByComponent.DAY.name,
            'report_type': GroupReportType.RECEIVED_DATA_REPORT.name,
        },
        empty_received_data_report_expected_file_contents,
    ),
    (
        {
            'group_by': GroupByComponent.MONTH.name,
            'report_type': GroupReportType.RECEIVED_DATA_REPORT.name,
        },
        empty_received_data_report_expected_file_contents,
    ),
    (
        {
            'group_by': GroupByComponent.YEAR.name,
            'report_type': GroupReportType.RECEIVED_DATA_REPORT.name,
        },
        empty_received_data_report_expected_file_contents,
    ),
    (
        {
            'group_by': GroupByComponent.DAY.name,
            'report_type': GroupReportType.APP_ACTIVITY_REPORT.name,
        },
        empty_app_activity_report_expected_file_contents,
    ),
    (
        {
            'group_by': GroupByComponent.MONTH.name,
            'report_type': GroupReportType.APP_ACTIVITY_REPORT.name,
        },
        empty_app_activity_report_expected_file_contents,
    ),
    (
        {
            'group_by': GroupByComponent.YEAR.name,
            'report_type': GroupReportType.APP_ACTIVITY_REPORT.name,
        },
        empty_app_activity_report_expected_file_contents,
    ),
]

# Content of empty individual summary reports
empty_individual_summary_report_expected_contents: list[dict[str, Any]] = [
    {
        'labs_summary_per_patient': [],
        'logins_summary_per_user': [],
        'patient_demographic_diagnosis_summary': [],
    },
]


def _create_filtered_group_report_stats_results() -> list[tuple[Any, Any]]:
    """Create filtered group report statistics results."""
    return [
        (
            {
                'group_by': GroupByComponent.DAY.name,
                'report_type': GroupReportType.SUMMARY_REPORT.name,
            },
            {
                'registration_summary': [
                    ['uncompleted_registration', 'completed_registration', 'total_registration_codes'],
                    ['2', '6', '8'],
                ],
                'grouped_registration_summary': [
                    ['day', 'uncompleted_registration', 'completed_registration', 'total_registration_codes'],
                    ['2025-01-08', '2', '6', '8'],
                ],
                'caregivers_summary': [
                    [
                        'caregivers_total',
                        'caregivers_registered',
                        'caregivers_unregistered',
                        'never_logged_in_after_registration',
                        'en',
                        'fr',
                    ],
                    ['6', '3', '3', '1', '2', '4'],
                ],
                'fetch_patients_summary': [
                    ['total', 'deceased', 'male', 'female', 'sex_other', 'sex_unknown', 'access_all', 'access_ntk'],
                    ['4', '0', '4', '0', '0', '0', '4', '0'],
                ],
                'devices_summary': [
                    ['device_total', 'device_ios', 'device_android', 'device_browser'],
                    ['6', '2', '2', '2'],
                ],
            },
        ),
        (
            {
                'group_by': GroupByComponent.MONTH.name,
                'report_type': GroupReportType.SUMMARY_REPORT.name,
            },
            {
                'registration_summary': [
                    ['uncompleted_registration', 'completed_registration', 'total_registration_codes'],
                    ['2', '6', '8'],
                ],
                'grouped_registration_summary': [
                    ['month', 'uncompleted_registration', 'completed_registration', 'total_registration_codes'],
                    ['2025-01-01', '2', '6', '8'],
                ],
                'caregivers_summary': [
                    [
                        'caregivers_total',
                        'caregivers_registered',
                        'caregivers_unregistered',
                        'never_logged_in_after_registration',
                        'en',
                        'fr',
                    ],
                    ['6', '3', '3', '1', '2', '4'],
                ],
                'fetch_patients_summary': [
                    ['total', 'deceased', 'male', 'female', 'sex_other', 'sex_unknown', 'access_all', 'access_ntk'],
                    ['4', '0', '4', '0', '0', '0', '4', '0'],
                ],
                'devices_summary': [
                    ['device_total', 'device_ios', 'device_android', 'device_browser'],
                    ['6', '2', '2', '2'],
                ],
            },
        ),
        (
            {
                'group_by': GroupByComponent.YEAR.name,
                'report_type': GroupReportType.SUMMARY_REPORT.name,
            },
            {
                'registration_summary': [
                    ['uncompleted_registration', 'completed_registration', 'total_registration_codes'],
                    ['2', '6', '8'],
                ],
                'grouped_registration_summary': [
                    ['year', 'uncompleted_registration', 'completed_registration', 'total_registration_codes'],
                    ['2025-01-01', '2', '6', '8'],
                ],
                'caregivers_summary': [
                    [
                        'caregivers_total',
                        'caregivers_registered',
                        'caregivers_unregistered',
                        'never_logged_in_after_registration',
                        'en',
                        'fr',
                    ],
                    ['6', '3', '3', '1', '2', '4'],
                ],
                'fetch_patients_summary': [
                    ['total', 'deceased', 'male', 'female', 'sex_other', 'sex_unknown', 'access_all', 'access_ntk'],
                    ['4', '0', '4', '0', '0', '0', '4', '0'],
                ],
                'devices_summary': [
                    ['device_total', 'device_ios', 'device_android', 'device_browser'],
                    ['6', '2', '2', '2'],
                ],
            },
        ),
        (
            {
                'group_by': GroupByComponent.DAY.name,
                'report_type': GroupReportType.RECEIVED_DATA_REPORT.name,
            },
            {
                'received_questionnaires_summary': [
                    [
                        'day',
                        'total_received_questionnaires',
                        'total_unique_patients',
                        'avg_received_questionnaires_per_patient',
                    ],
                    ['2025-01-08', '30', '3', '10'],
                    ['2025-01-07', '5', '1', '5'],
                    ['2025-01-06', '15', '1', '15'],
                ],
                'received_documents_summary': [
                    ['day', 'total_received_documents', 'total_unique_patients', 'avg_received_documents_per_patient'],
                    ['2025-01-08', '30', '3', '10'],
                    ['2025-01-07', '5', '1', '5'],
                    ['2025-01-06', '15', '1', '15'],
                ],
                'received_educational_materials_summary': [
                    [
                        'day',
                        'total_received_edu_materials',
                        'total_unique_patients',
                        'avg_received_edu_materials_per_patient',
                    ],
                    ['2025-01-08', '30', '3', '10'],
                    ['2025-01-07', '5', '1', '5'],
                    ['2025-01-06', '15', '1', '15'],
                ],
                'received_appointments_summary': [
                    [
                        'day',
                        'total_received_appointments',
                        'total_unique_patients',
                        'avg_received_appointments_per_patient',
                    ],
                    ['2025-01-08', '30', '3', '10'],
                    ['2025-01-07', '5', '1', '5'],
                    ['2025-01-06', '15', '1', '15'],
                ],
                'received_labs_summary': [
                    ['day', 'total_received_labs', 'total_unique_patients', 'avg_received_labs_per_patient'],
                    ['2025-01-08', '30', '3', '10'],
                    ['2025-01-07', '15', '1', '15'],
                    ['2025-01-06', '15', '1', '15'],
                ],
                'received_clinical_data_summary': [
                    [
                        'no_appointments_labs_notes',
                        'has_appointments_only',
                        'has_labs_only',
                        'has_clinical_notes_only',
                        'receiving_new_data_total',
                    ],
                    ['0', '0', '0', '0', '5'],
                ],
            },
        ),
        (
            {
                'group_by': GroupByComponent.MONTH.name,
                'report_type': GroupReportType.RECEIVED_DATA_REPORT.name,
            },
            {
                'received_questionnaires_summary': [
                    [
                        'month',
                        'total_received_questionnaires',
                        'total_unique_patients',
                        'avg_received_questionnaires_per_patient',
                    ],
                    ['2025-01-01', '50', '4', '12'],
                ],
                'received_documents_summary': [
                    [
                        'month',
                        'total_received_documents',
                        'total_unique_patients',
                        'avg_received_documents_per_patient',
                    ],
                    ['2025-01-01', '50', '4', '12'],
                ],
                'received_educational_materials_summary': [
                    [
                        'month',
                        'total_received_edu_materials',
                        'total_unique_patients',
                        'avg_received_edu_materials_per_patient',
                    ],
                    ['2025-01-01', '50', '4', '12'],
                ],
                'received_appointments_summary': [
                    [
                        'month',
                        'total_received_appointments',
                        'total_unique_patients',
                        'avg_received_appointments_per_patient',
                    ],
                    ['2025-01-01', '50', '4', '12'],
                ],
                'received_labs_summary': [
                    ['month', 'total_received_labs', 'total_unique_patients', 'avg_received_labs_per_patient'],
                    ['2025-01-01', '60', '4', '15'],
                ],
                'received_clinical_data_summary': [
                    [
                        'no_appointments_labs_notes',
                        'has_appointments_only',
                        'has_labs_only',
                        'has_clinical_notes_only',
                        'receiving_new_data_total',
                    ],
                    ['0', '0', '0', '0', '5'],
                ],
            },
        ),
        (
            {
                'group_by': GroupByComponent.YEAR.name,
                'report_type': GroupReportType.RECEIVED_DATA_REPORT.name,
            },
            {
                'received_questionnaires_summary': [
                    [
                        'year',
                        'total_received_questionnaires',
                        'total_unique_patients',
                        'avg_received_questionnaires_per_patient',
                    ],
                    ['2025-01-01', '50', '4', '12'],
                ],
                'received_documents_summary': [
                    [
                        'year',
                        'total_received_documents',
                        'total_unique_patients',
                        'avg_received_documents_per_patient',
                    ],
                    ['2025-01-01', '50', '4', '12'],
                ],
                'received_educational_materials_summary': [
                    [
                        'year',
                        'total_received_edu_materials',
                        'total_unique_patients',
                        'avg_received_edu_materials_per_patient',
                    ],
                    ['2025-01-01', '50', '4', '12'],
                ],
                'received_appointments_summary': [
                    [
                        'year',
                        'total_received_appointments',
                        'total_unique_patients',
                        'avg_received_appointments_per_patient',
                    ],
                    ['2025-01-01', '50', '4', '12'],
                ],
                'received_labs_summary': [
                    ['year', 'total_received_labs', 'total_unique_patients', 'avg_received_labs_per_patient'],
                    ['2025-01-01', '60', '4', '15'],
                ],
                'received_clinical_data_summary': [
                    [
                        'no_appointments_labs_notes',
                        'has_appointments_only',
                        'has_labs_only',
                        'has_clinical_notes_only',
                        'receiving_new_data_total',
                    ],
                    ['0', '0', '0', '0', '5'],
                ],
            },
        ),
        (
            {
                'group_by': GroupByComponent.DAY.name,
                'report_type': GroupReportType.APP_ACTIVITY_REPORT.name,
            },
            {
                'logins_summary': [
                    ['day', 'total_logins', 'unique_user_logins', 'avg_logins_per_user'],
                    ['2025-01-08', '18', '3', '6'],
                    ['2025-01-06', '8', '2', '4'],
                ],
                'users_clicks_summary': [
                    ['day', 'login_count', 'feedback_count', 'update_security_answers_count', 'update_passwords_count'],
                    ['2025-01-08', '18', '21', '24', '27'],
                    ['2025-01-06', '8', '10', '12', '14'],
                ],
                'user_patient_clicks_summary': [
                    [
                        'day',
                        'checkins_count',
                        'documents_count',
                        'educational_materials_count',
                        'completed_questionnaires_count',
                        'labs_count',
                    ],
                    ['2025-01-08', '4', '6', '8', '10', '12'],
                    ['2025-01-07', '16', '20', '24', '28', '32'],
                    ['2025-01-06', '13', '15', '17', '19', '21'],
                ],
                'users_latest_login_year_summary': [
                    ['2025'],
                    ['3'],
                ],
            },
        ),
        (
            {
                'group_by': GroupByComponent.MONTH.name,
                'report_type': GroupReportType.APP_ACTIVITY_REPORT.name,
            },
            {
                'logins_summary': [
                    ['month', 'total_logins', 'unique_user_logins', 'avg_logins_per_user'],
                    ['2025-01-01', '26', '3', '8'],
                ],
                'users_clicks_summary': [
                    [
                        'month',
                        'login_count',
                        'feedback_count',
                        'update_security_answers_count',
                        'update_passwords_count',
                    ],
                    ['2025-01-01', '26', '31', '36', '41'],
                ],
                'user_patient_clicks_summary': [
                    [
                        'month',
                        'checkins_count',
                        'documents_count',
                        'educational_materials_count',
                        'completed_questionnaires_count',
                        'labs_count',
                    ],
                    ['2025-01-01', '33', '41', '49', '57', '65'],
                ],
                'users_latest_login_year_summary': [
                    ['2025'],
                    ['3'],
                ],
            },
        ),
        (
            {
                'group_by': GroupByComponent.YEAR.name,
                'report_type': GroupReportType.APP_ACTIVITY_REPORT.name,
            },
            {
                'logins_summary': [
                    ['year', 'total_logins', 'unique_user_logins', 'avg_logins_per_user'],
                    ['2025-01-01', '26', '3', '8'],
                ],
                'users_clicks_summary': [
                    [
                        'year',
                        'login_count',
                        'feedback_count',
                        'update_security_answers_count',
                        'update_passwords_count',
                    ],
                    ['2025-01-01', '26', '31', '36', '41'],
                ],
                'user_patient_clicks_summary': [
                    [
                        'year',
                        'checkins_count',
                        'documents_count',
                        'educational_materials_count',
                        'completed_questionnaires_count',
                        'labs_count',
                    ],
                    [
                        '2025-01-01',
                        '33',
                        '41',
                        '49',
                        '57',
                        '65',
                    ],
                ],
                'users_latest_login_year_summary': [
                    ['2025'],
                    ['3'],
                ],
            },
        ),
    ]


def _create_individual_summary_report_expected_contents() -> dict[str, Any]:
    """Create expected individual summary report statistics."""
    return {
        'labs_summary_per_patient': [
            [
                'patient__legacy_id',
                'patient_ser_num',
                'first_lab_received_utc',
                'last_lab_received_utc',
                'total_labs_received',
            ],
            ['51', '51', '2024-12-31 17:00:00+00:00', '2025-01-08 17:00:00+00:00', '20'],
            ['52', '52', '2025-01-08 17:00:00+00:00', '2025-01-08 17:00:00+00:00', '10'],
            ['53', '53', '2025-01-08 17:00:00+00:00', '2025-01-08 17:00:00+00:00', '15'],
            ['54', '54', '2025-01-08 17:00:00+00:00', '2025-01-08 17:00:00+00:00', '15'],
        ],
        'logins_summary_per_user': [
            ['user_id', 'total_logged_in_days', 'total_logins', 'avg_logins_per_day'],
            ['1', '2', '13', '6'],
            ['2', '2', '8', '4'],
            ['3', '1', '5', '5'],
            ['4', '1', '3', '3'],
            ['5', '1', '5', '5'],
            ['6', '1', '5', '5'],
        ],
        'patient_demographic_diagnosis_summary': [
            [
                'patient_ser_num',
                'age',
                'date_of_birth',
                'sex',
                'email',
                'language',
                'registration_date_utc',
                'latest_diagnosis_description',
                'latest_diagnosis_date_utc',
            ],
            [
                '51',
                '',
                '2018-01-01 05:00:00+00:00',
                'Male',
                'test@test.com',
                'EN',
                '2025-01-08 17:00:00+00:00',
                'Test Diagnosis2',
                '2025-01-08 17:00:00+00:00',
            ],
            [
                '52',
                '',
                '2018-01-01 05:00:00+00:00',
                'Male',
                'test@test.com',
                'EN',
                '2025-01-08 17:00:00+00:00',
                'Test Diagnosis1',
                '2025-01-08 17:00:00+00:00',
            ],
            [
                '53',
                '',
                '2018-01-01 05:00:00+00:00',
                'Male',
                'test@test.com',
                'EN',
                '2025-01-08 17:00:00+00:00',
                'Test Diagnosis1',
                '2025-01-08 17:00:00+00:00',
            ],
            ['54', '', '2018-01-01 05:00:00+00:00', 'Male', 'test@test.com', 'EN', '2025-01-08 17:00:00+00:00', '', ''],
        ],
    }


# Add any future GET-requestable usage statistics pages here for faster test writing
test_url_template_data: list[tuple[str, str]] = [
    (reverse('usage-statistics:reports-group-export'), 'usage_statistics/reports/export_form.html'),
    (reverse('usage-statistics:reports-individual-export'), 'usage_statistics/reports/export_form.html'),
]


@pytest.mark.parametrize(('url', 'template'), test_url_template_data)
def test_usage_statistics_urls_exist(admin_client: Client, url: str, template: str) -> None:
    """Ensure that a page exists at each URL address."""
    response = admin_client.get(url)

    assert response.status_code == HTTPStatus.OK


@pytest.mark.parametrize(('url', 'template'), test_url_template_data)
def test_statistics_views_use_correct_template(admin_client: Client, url: str, template: str) -> None:
    """Ensure that a page uses appropriate templates."""
    response = admin_client.get(url)

    assertTemplateUsed(response, template)


def test_usage_statistics_group_export_unauthorized(user_client: Client) -> None:
    """Ensure that an authenticated (not admin) user cannot access the group reports page."""
    response = user_client.get(reverse('usage-statistics:reports-group-export'))

    assertContains(
        response=response,
        text='403 Forbidden',
        status_code=HTTPStatus.FORBIDDEN,
    )


def test_usage_statistics_individual_export_unauthorized(user_client: Client) -> None:
    """Ensure that an authenticated (not admin) user cannot access the individual reports page."""
    response = user_client.get(reverse('usage-statistics:reports-individual-export'))

    assertContains(
        response=response,
        text='403 Forbidden',
        status_code=HTTPStatus.FORBIDDEN,
    )


@pytest.mark.parametrize(('report_filter', 'expected_csv_content'), empty_group_report_filters_data)
def test_usage_stats_group_reports_empty_csvs(
    client: Client,
    admin_user: User,
    report_filter: dict[str, Any],
    expected_csv_content: dict[str, Any],
) -> None:
    """Ensure that group report request successfully returns empty reports in CSV format."""
    client.force_login(user=admin_user)

    url = reverse('usage-statistics:reports-group-export')

    group_usage_stats_form = GroupUsageStatisticsForm(
        data={
            'start_date': timezone.now().date() - dt.timedelta(days=7),
            'end_date': timezone.now().date(),
        }
        | report_filter,
    )

    assert group_usage_stats_form.is_valid()

    response = client.post(
        url,
        data=group_usage_stats_form.cleaned_data | {'download_csv': ['Download CSV']},
    )

    assert response.status_code == HTTPStatus.OK
    assert isinstance(response, FileResponse)
    assert response['Content-Type'] == 'application/zip'

    # Collect chunks from the streaming_content into a buffer
    zip_buffer = io.BytesIO()
    # Cast to Iterator[bytes] to ensure that streaming_content is being iterated in a synchronous manner
    streaming_content = cast('Iterator[bytes]', response.streaming_content)
    for chunk in streaming_content:
        zip_buffer.write(chunk)
    zip_buffer.seek(0)

    with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
        actual_files = set(zip_file.namelist())
        expected_files = {f'{key}.csv' for key in expected_csv_content}
        missing_files = expected_files - actual_files
        assert not missing_files, (f'Missing expected files in the ZIP: {missing_files}',)

        # For each expected file, check its contents
        for filename, expected_content in expected_csv_content.items():
            with zip_file.open(f'{filename}.csv', 'r') as file_in_zip:
                file_content = file_in_zip.read().decode(errors='replace')
                # Filter out empty or whitespace-only lines before creating the reader
                filtered_content = [line for line in file_content.splitlines() if line.strip()]
                reader = csv.reader(filtered_content)

                rows = list(reader)

                assert rows == expected_content, (
                    f'Mismatch in {filename} contents:\n' + f'  Expected: {expected_content}\n' + f'  Actual:   {rows}'
                )


@pytest.mark.parametrize(('report_filter', 'expected_csv_content'), _create_filtered_group_report_stats_results())
def test_usage_stats_group_reports_csvs_with_data(
    client: Client,
    admin_user: User,
    mocker: MockerFixture,
    report_filter: dict[str, Any],
    expected_csv_content: dict[str, Any],
) -> None:
    """Ensure that group report request successfully returns reports with statistics in CSV format."""
    now = dt.datetime(2025, 1, 8, 12, 0, 0, tzinfo=timezone.get_current_timezone())
    mocker.patch('django.utils.timezone.now', return_value=now)
    _create_received_medical_records()
    _create_device_identifier_records()
    _create_registration_records(mocker)
    mocker.patch('django.utils.timezone.now', return_value=now)
    _create_app_activity_records()
    _create_user_patient_activity_records()

    client.force_login(user=admin_user)

    url = reverse('usage-statistics:reports-group-export')

    group_usage_stats_form = GroupUsageStatisticsForm(
        data={
            'start_date': now.date() - dt.timedelta(days=7),
            'end_date': now.date(),
        }
        | report_filter,
    )

    assert group_usage_stats_form.is_valid()

    response = client.post(
        url,
        data=group_usage_stats_form.cleaned_data | {'download_csv': ['Download CSV']},
    )

    assert response.status_code == HTTPStatus.OK
    assert isinstance(response, FileResponse)
    assert response['Content-Type'] == 'application/zip'

    # Collect chunks from the streaming_content into a buffer
    zip_buffer = io.BytesIO()
    # Cast to Iterator[bytes] to ensure that streaming_content is being iterated in a synchronous manner
    streaming_content = cast('Iterator[bytes]', response.streaming_content)
    for chunk in streaming_content:
        zip_buffer.write(chunk)
    zip_buffer.seek(0)

    with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
        actual_files = set(zip_file.namelist())
        expected_files = {f'{key}.csv' for key in expected_csv_content}
        missing_files = expected_files - actual_files
        assert not missing_files, (f'Missing expected files in the ZIP: {missing_files}',)

        # For each expected file, check its contents
        for filename, expected_content in expected_csv_content.items():
            with zip_file.open(f'{filename}.csv', 'r') as file_in_zip:
                file_content = file_in_zip.read().decode(errors='replace')
                # Filter out empty or whitespace-only lines before creating the reader
                filtered_content = [line for line in file_content.splitlines() if line.strip()]
                reader = csv.reader(filtered_content)

                rows = list(reader)

                assert rows == expected_content, (
                    f'Mismatch in {filename} contents:\n' + f'  Expected: {expected_content}\n' + f'  Actual:   {rows}'
                )


def test_usage_stats_group_reports_bad_request(
    client: Client,
    admin_user: User,
) -> None:
    """Ensure that group report request returns an HttpResponseBadRequest if no valid download option is selected."""
    client.force_login(user=admin_user)

    url = reverse('usage-statistics:reports-group-export')

    group_usage_stats_form = GroupUsageStatisticsForm(
        data={
            'start_date': timezone.now().date() - dt.timedelta(days=7),
            'end_date': timezone.now().date(),
            'group_by': GroupByComponent.YEAR.name,
            'report_type': GroupReportType.SUMMARY_REPORT.name,
        },
    )

    assert group_usage_stats_form.is_valid()

    response = client.post(
        url,
        data=group_usage_stats_form.cleaned_data,
    )

    assert response.status_code == HTTPStatus.BAD_REQUEST
    assert isinstance(response, HttpResponseBadRequest)


@pytest.mark.parametrize(('report_filter', 'expected_xlsx_content'), empty_group_report_filters_data)
def test_usage_stats_group_reports_empty_xlsx(
    client: Client,
    admin_user: User,
    report_filter: dict[str, Any],
    expected_xlsx_content: dict[str, Any],
) -> None:
    """Ensure that group report request successfully returns empty reports in XLSX format."""
    client.force_login(user=admin_user)

    url = reverse('usage-statistics:reports-group-export')

    group_usage_stats_form = GroupUsageStatisticsForm(
        data={
            'start_date': timezone.now().date() - dt.timedelta(days=7),
            'end_date': timezone.now().date(),
        }
        | report_filter,
    )

    assert group_usage_stats_form.is_valid()

    response = client.post(
        url,
        data=group_usage_stats_form.cleaned_data | {'download_xlsx': ['Download XLSX']},
    )

    assert response.status_code == HTTPStatus.OK
    assert isinstance(response, FileResponse)
    assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Collect chunks from the streaming_content into a buffer
    xlsx_buffer = io.BytesIO()
    # Cast to Iterator[bytes] to ensure that streaming_content is being iterated in a synchronous manner
    streaming_content = cast('Iterator[bytes]', response.streaming_content)
    for chunk in streaming_content:
        xlsx_buffer.write(chunk)

    workbook = load_workbook(xlsx_buffer)

    # For each expected worksheet, verify presence and the contents
    for sheet_name, expected_content in expected_xlsx_content.items():
        truncator = Truncator(sheet_name)
        truncated_sheet_name = truncator.chars(num=utils.SHEET_TITLE_MAX_LENGTH)

        assert truncated_sheet_name in workbook.sheetnames, (f'Workbook missing expected sheet {truncated_sheet_name}',)

        sheet = workbook[truncated_sheet_name]

        actual_rows = []
        for row in sheet.iter_rows():
            # Convert each cell value to a string (None -> "")
            row_values = [(str(cell.value) if cell.value is not None else '') for cell in row]

            actual_rows.append(row_values)

        assert actual_rows == expected_content, (
            f'Mismatch in {sheet_name} contents:\n' + f'  Expected: {expected_content}\n' + f'  Actual:   {actual_rows}'
        )


@pytest.mark.parametrize(('report_filter', 'expected_xlsx_content'), _create_filtered_group_report_stats_results())
def test_usage_stats_group_reports_xlsx_with_data(
    client: Client,
    admin_user: User,
    mocker: MockerFixture,
    report_filter: dict[str, Any],
    expected_xlsx_content: dict[str, Any],
) -> None:
    """Ensure that group report request successfully returns reports with statistics in XLSX format."""
    now = dt.datetime(2025, 1, 8, 12, 0, 0, tzinfo=timezone.get_current_timezone())
    mocker.patch('django.utils.timezone.now', return_value=now)
    _create_received_medical_records()
    _create_device_identifier_records()
    _create_registration_records(mocker)
    mocker.patch('django.utils.timezone.now', return_value=now)
    _create_app_activity_records()
    _create_user_patient_activity_records()

    client.force_login(user=admin_user)

    url = reverse('usage-statistics:reports-group-export')

    group_usage_stats_form = GroupUsageStatisticsForm(
        data={
            'start_date': now.date() - dt.timedelta(days=7),
            'end_date': now.date(),
        }
        | report_filter,
    )

    assert group_usage_stats_form.is_valid()

    response = client.post(
        url,
        data=group_usage_stats_form.cleaned_data | {'download_xlsx': ['Download XLSX']},
    )

    assert response.status_code == HTTPStatus.OK
    assert isinstance(response, FileResponse)
    assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Collect chunks from the streaming_content into a buffer
    xlsx_buffer = io.BytesIO()
    # Cast to Iterator[bytes] to ensure that streaming_content is being iterated in a synchronous manner
    streaming_content = cast('Iterator[bytes]', response.streaming_content)
    for chunk in streaming_content:
        xlsx_buffer.write(chunk)

    workbook = load_workbook(xlsx_buffer)

    # For each expected worksheet, verify presence and the contents
    for sheet_name, expected_content in expected_xlsx_content.items():
        if sheet_name in {
            'received_questionnaires_summary',
            'received_documents_summary',
            'received_educational_materials_summary',
            'received_appointments_summary',
            'received_labs_summary',
            'logins_summary',
            'users_clicks_summary',
            'user_patient_clicks_summary',
            'grouped_registration_summary',
        }:
            max_range = len(expected_content)
            for item in range(1, max_range):
                expected_content[item][0] = f'{expected_content[item][0]} 00:00:00'

        truncator = Truncator(sheet_name)
        truncated_sheet_name = truncator.chars(num=utils.SHEET_TITLE_MAX_LENGTH)

        assert truncated_sheet_name in workbook.sheetnames, (f'Workbook missing expected sheet {truncated_sheet_name}',)

        sheet = workbook[truncated_sheet_name]

        actual_rows = []
        for row in sheet.iter_rows():
            # Convert each cell value to a string (None -> "")
            row_values = [(str(cell.value) if cell.value is not None else '') for cell in row]

            actual_rows.append(row_values)

        assert actual_rows == expected_content, (
            f'Mismatch in {sheet_name} contents:\n' + f'  Expected: {expected_content}\n' + f'  Actual:   {actual_rows}'
        )


@pytest.mark.parametrize('expected_csv_content', empty_individual_summary_report_expected_contents)
def test_usage_stats_individual_reports_empty_csvs(
    client: Client,
    admin_user: User,
    expected_csv_content: dict[str, Any],
) -> None:
    """Ensure that individual report request successfully returns empty reports in CSV format."""
    client.force_login(user=admin_user)

    url = reverse('usage-statistics:reports-individual-export')

    individual_usage_stats_form = IndividualUsageStatisticsForm(data={})

    assert individual_usage_stats_form.is_valid()

    response = client.post(
        url,
        data=individual_usage_stats_form.cleaned_data | {'download_csv': ['Download CSV']},
    )

    assert response.status_code == HTTPStatus.OK
    assert isinstance(response, FileResponse)
    assert response['Content-Type'] == 'application/zip'

    # Collect chunks from the streaming_content into a buffer
    zip_buffer = io.BytesIO()
    # Cast to Iterator[bytes] to ensure that streaming_content is being iterated in a synchronous manner
    streaming_content = cast('Iterator[bytes]', response.streaming_content)
    for chunk in streaming_content:
        zip_buffer.write(chunk)
    zip_buffer.seek(0)

    with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
        actual_files = set(zip_file.namelist())
        expected_files = {f'{key}.csv' for key in expected_csv_content}
        missing_files = expected_files - actual_files
        assert not missing_files, (f'Missing expected files in the ZIP: {missing_files}',)

        # For each expected file, check its contents
        for filename, expected_content in expected_csv_content.items():
            with zip_file.open(f'{filename}.csv', 'r') as file_in_zip:
                file_content = file_in_zip.read().decode(errors='replace')
                # Filter out empty or whitespace-only lines before creating the reader
                filtered_content = [line for line in file_content.splitlines() if line.strip()]
                reader = csv.reader(filtered_content)

                rows = list(reader)

                assert rows == expected_content, (
                    f'Mismatch in {filename} contents:\n' + f'  Expected: {expected_content}\n' + f'  Actual:   {rows}'
                )


def test_usage_stats_individual_reports_expected_csvs(
    client: Client,
    admin_user: User,
    mocker: MockerFixture,
) -> None:
    """Ensure that individual report request successfully returns expected reports in CSV format."""
    now = dt.datetime(2025, 1, 8, 12, 0, 0, tzinfo=timezone.get_current_timezone())
    mocker.patch('django.utils.timezone.now', return_value=now)
    _create_received_medical_records()
    _create_registration_records(mocker)
    mocker.patch('django.utils.timezone.now', return_value=now)
    _create_app_activity_records()
    _create_demographic_diagnosis_records()

    client.force_login(user=admin_user)

    url = reverse('usage-statistics:reports-individual-export')

    individual_usage_stats_form = IndividualUsageStatisticsForm(data={})

    assert individual_usage_stats_form.is_valid()

    response = client.post(
        url,
        data=individual_usage_stats_form.cleaned_data | {'download_csv': ['Download CSV']},
    )

    assert response.status_code == HTTPStatus.OK
    assert isinstance(response, FileResponse)
    assert response['Content-Type'] == 'application/zip'

    # Collect chunks from the streaming_content into a buffer
    zip_buffer = io.BytesIO()
    # Cast to Iterator[bytes] to ensure that streaming_content is being iterated in a synchronous manner
    streaming_content = cast('Iterator[bytes]', response.streaming_content)
    for chunk in streaming_content:
        zip_buffer.write(chunk)
    zip_buffer.seek(0)

    individual_summary_report_expected_contents = _create_individual_summary_report_expected_contents()
    # Ensure that expected user IDs are the same as in the database
    individual_summary_report_expected_contents['logins_summary_per_user'][1][0] = str(
        Caregiver.objects.filter(username='marge')[0].id,
    )
    individual_summary_report_expected_contents['logins_summary_per_user'][2][0] = str(
        Caregiver.objects.filter(username='homer')[0].id,
    )
    individual_summary_report_expected_contents['logins_summary_per_user'][3][0] = str(
        Caregiver.objects.filter(username='bart')[0].id,
    )
    individual_summary_report_expected_contents['logins_summary_per_user'][4][0] = str(
        Caregiver.objects.filter(username='lisa')[0].id,
    )
    individual_summary_report_expected_contents['logins_summary_per_user'][5][0] = str(
        Caregiver.objects.filter(username='mona')[0].id,
    )
    individual_summary_report_expected_contents['logins_summary_per_user'][6][0] = str(
        Caregiver.objects.filter(username='fred')[0].id,
    )

    with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
        actual_files = set(zip_file.namelist())
        expected_files = {f'{key}.csv' for key in individual_summary_report_expected_contents}
        missing_files = expected_files - actual_files
        assert not missing_files, (f'Missing expected files in the ZIP: {missing_files}',)

        # For each expected file, check its contents
        for filename, expected_content in individual_summary_report_expected_contents.items():
            with zip_file.open(f'{filename}.csv', 'r') as file_in_zip:
                file_content = file_in_zip.read().decode(errors='replace')
                # Filter out empty or whitespace-only lines before creating the reader
                filtered_content = [line for line in file_content.splitlines() if line.strip()]
                reader = csv.reader(filtered_content)

                rows = list(reader)

                assert rows == expected_content, (
                    f'Mismatch in {filename} contents:\n' + f'  Expected: {expected_content}\n' + f'  Actual:   {rows}'
                )


@pytest.mark.parametrize('expected_xlsx_content', empty_individual_summary_report_expected_contents)
def test_usage_stats_individual_reports_empty_xlsx(
    client: Client,
    admin_user: User,
    expected_xlsx_content: dict[str, Any],
) -> None:
    """Ensure that individual report request successfully returns empty reports in XLSX format."""
    client.force_login(user=admin_user)

    url = reverse('usage-statistics:reports-individual-export')

    group_usage_stats_form = IndividualUsageStatisticsForm(data={})

    assert group_usage_stats_form.is_valid()

    response = client.post(
        url,
        data=group_usage_stats_form.cleaned_data | {'download_xlsx': ['Download XLSX']},
    )

    assert response.status_code == HTTPStatus.OK
    assert isinstance(response, FileResponse)
    assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Collect chunks from the streaming_content into a buffer
    xlsx_buffer = io.BytesIO()
    # Cast to Iterator[bytes] to ensure that streaming_content is being iterated in a synchronous manner
    streaming_content = cast('Iterator[bytes]', response.streaming_content)
    for chunk in streaming_content:
        xlsx_buffer.write(chunk)

    workbook = load_workbook(xlsx_buffer)

    # For each expected worksheet, verify presence and the contents
    for sheet_name, expected_content in expected_xlsx_content.items():
        truncator = Truncator(sheet_name)
        truncated_sheet_name = truncator.chars(num=utils.SHEET_TITLE_MAX_LENGTH)

        assert truncated_sheet_name in workbook.sheetnames, (f'Workbook missing expected sheet {truncated_sheet_name}',)

        sheet = workbook[truncated_sheet_name]

        actual_rows = []
        for row in sheet.iter_rows():
            # Convert each cell value to a string (None -> "")
            row_values = [(str(cell.value) if cell.value is not None else '') for cell in row]

            actual_rows.append(row_values)

        assert actual_rows == expected_content, (
            f'Mismatch in {sheet_name} contents:\n' + f'  Expected: {expected_content}\n' + f'  Actual:   {actual_rows}'
        )


def test_usage_stats_individual_reports_expected_xlsx_data(
    client: Client,
    admin_user: User,
    mocker: MockerFixture,
) -> None:
    """Ensure that group report request successfully returns reports with statistics in XLSX format."""
    now = dt.datetime(2025, 1, 8, 12, 0, 0, tzinfo=timezone.get_current_timezone())
    mocker.patch('django.utils.timezone.now', return_value=now)
    _create_received_medical_records()
    _create_registration_records(mocker)
    mocker.patch('django.utils.timezone.now', return_value=now)
    _create_app_activity_records()
    _create_demographic_diagnosis_records()

    client.force_login(user=admin_user)

    url = reverse('usage-statistics:reports-individual-export')

    individual_usage_stats_form = IndividualUsageStatisticsForm(data={})

    assert individual_usage_stats_form.is_valid()

    response = client.post(
        url,
        data=individual_usage_stats_form.cleaned_data | {'download_xlsx': ['Download XLSX']},
    )

    assert response.status_code == HTTPStatus.OK
    assert isinstance(response, FileResponse)
    assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    individual_summary_report_expected_contents = _create_individual_summary_report_expected_contents()
    # Ensure that expected user IDs are the same as in the database
    individual_summary_report_expected_contents['logins_summary_per_user'][1][0] = str(
        Caregiver.objects.filter(username='marge')[0].id,
    )
    individual_summary_report_expected_contents['logins_summary_per_user'][2][0] = str(
        Caregiver.objects.filter(username='homer')[0].id,
    )
    individual_summary_report_expected_contents['logins_summary_per_user'][3][0] = str(
        Caregiver.objects.filter(username='bart')[0].id,
    )
    individual_summary_report_expected_contents['logins_summary_per_user'][4][0] = str(
        Caregiver.objects.filter(username='lisa')[0].id,
    )
    individual_summary_report_expected_contents['logins_summary_per_user'][5][0] = str(
        Caregiver.objects.filter(username='mona')[0].id,
    )
    individual_summary_report_expected_contents['logins_summary_per_user'][6][0] = str(
        Caregiver.objects.filter(username='fred')[0].id,
    )

    # Collect chunks from the streaming_content into a buffer
    xlsx_buffer = io.BytesIO()
    # Cast to Iterator[bytes] to ensure that streaming_content is being iterated in a synchronous manner
    streaming_content = cast('Iterator[bytes]', response.streaming_content)
    for chunk in streaming_content:
        xlsx_buffer.write(chunk)

    workbook = load_workbook(xlsx_buffer)

    # For each expected worksheet, verify presence and the contents
    for sheet_name, expected_content in individual_summary_report_expected_contents.items():
        if sheet_name == 'labs_summary_per_patient':
            for lab_summary in expected_content[1:]:
                lab_summary[2] = lab_summary[2][:-6]
                lab_summary[3] = lab_summary[3][:-6]
        if sheet_name == 'patient_demographic_diagnosis_summary':
            for patient_summary_content in expected_content[1:]:
                patient_summary_content[2] = patient_summary_content[2][:-6]
                patient_summary_content[6] = patient_summary_content[6][:-6]
                patient_summary_content[8] = patient_summary_content[8][:-6]

        truncator = Truncator(sheet_name)
        truncated_sheet_name = truncator.chars(num=utils.SHEET_TITLE_MAX_LENGTH)

        assert truncated_sheet_name in workbook.sheetnames, (f'Workbook missing expected sheet {truncated_sheet_name}',)

        sheet = workbook[truncated_sheet_name]

        actual_rows = []
        for row in sheet.iter_rows():
            # Convert each cell value to a string (None -> "")
            row_values = [(str(cell.value) if cell.value is not None else '') for cell in row]

            actual_rows.append(row_values)

        assert actual_rows == expected_content, (
            f'Mismatch in {sheet_name} contents:\n' + f'  Expected: {expected_content}\n' + f'  Actual:   {actual_rows}'
        )


def _create_registration_records(mocker: MockerFixture) -> None:
    """
    Create registration records for 4 patients.

    The records are created for Marge, Homer, Bart, and Lisa.
    """
    relationships = _create_relationship_records()

    caregiver_models.RegistrationCode.objects.bulk_create([
        caregiver_models.RegistrationCode(
            code='marge_code',
            relationship=relationships['marge_relationship'],
            status=caregiver_models.RegistrationCodeStatus.REGISTERED,
        ),
        caregiver_models.RegistrationCode(
            code='marge_homer',
            relationship=relationships['marge_homer_relationship'],
            status=caregiver_models.RegistrationCodeStatus.REGISTERED,
        ),
        caregiver_models.RegistrationCode(
            code='homer_self1',
            relationship=relationships['homer_relationship'],
            status=caregiver_models.RegistrationCodeStatus.BLOCKED,
        ),
        caregiver_models.RegistrationCode(
            code='homer_self2',
            relationship=relationships['homer_pending_relationship'],
            status=caregiver_models.RegistrationCodeStatus.REGISTERED,
        ),
        caregiver_models.RegistrationCode(
            code='marge_bart',
            relationship=relationships['marge_bart_relationship'],
            status=caregiver_models.RegistrationCodeStatus.REGISTERED,
        ),
        caregiver_models.RegistrationCode(
            code='bart_self',
            relationship=relationships['bart_relationship'],
            status=caregiver_models.RegistrationCodeStatus.REGISTERED,
        ),
        caregiver_models.RegistrationCode(
            code='homer_lisa',
            relationship=relationships['homer_lisa_relationship'],
            status=caregiver_models.RegistrationCodeStatus.REGISTERED,
        ),
        caregiver_models.RegistrationCode(
            code='lisa_self2',
            relationship=relationships['lisa_pending_relationship'],
            status=caregiver_models.RegistrationCodeStatus.NEW,
        ),
    ])
    # Lisa's registration code created a month ago
    previous_month = timezone.now() - dt.timedelta(days=31)
    mocker.patch('django.utils.timezone.now', return_value=previous_month)
    caregiver_factories.RegistrationCode.create(
        code='lisa_self1',
        relationship=relationships['lisa_relationship'],
        status=caregiver_models.RegistrationCodeStatus.REGISTERED,
    )


def _create_relationship_records() -> dict[str, Any]:
    """
    Create relationships for 4 patients.

    The records are created for Marge, Homer, Bart, and Lisa.

    Returns:
        dictionary with self relationships
    """
    marge_caregiver = user_factories.Caregiver.create(
        username='marge',
        language='fr',
        last_login=timezone.now(),
        date_joined=timezone.now(),
    )
    homer_caregiver = user_factories.Caregiver.create(
        username='homer',
        language='fr',
        last_login=timezone.now(),
        date_joined=timezone.now(),
    )
    bart_caregiver = user_factories.Caregiver.create(username='bart', date_joined=timezone.now())
    lisa_caregiver = user_factories.Caregiver.create(
        username='lisa',
        is_active=False,
        date_joined=timezone.now(),
    )
    user_factories.Caregiver.create(
        username='mona',
        language='fr',
        is_active=False,
        date_joined=timezone.now(),
    )
    user_factories.Caregiver.create(
        username='fred',
        language='fr',
        is_active=False,
        date_joined=timezone.now(),
    )

    marge_caregiver_profile = caregiver_factories.CaregiverProfile.create(
        user=marge_caregiver,
        legacy_id=1,
    )
    homer_caregiver_profile = caregiver_factories.CaregiverProfile.create(
        user=homer_caregiver,
        legacy_id=2,
    )
    bart_caregiver_profile = caregiver_factories.CaregiverProfile.create(user=bart_caregiver, legacy_id=3)

    lisa_caregiver_profile = caregiver_factories.CaregiverProfile.create(user=lisa_caregiver, legacy_id=4)

    marge_patient = patient_factories.Patient.create(legacy_id=51, ramq='TEST01161972')
    homer_patient = patient_factories.Patient.create(legacy_id=52, ramq='TEST01161973')
    bart_patient = patient_factories.Patient.create(legacy_id=53, ramq='TEST01161974')
    lisa_patient = patient_factories.Patient.create(legacy_id=54, ramq='TEST01161975')

    # marge
    marge_self_relationship = patient_factories.Relationship.create(
        type=patient_models.RelationshipType.objects.self_type(),
        patient=marge_patient,
        caregiver=marge_caregiver_profile,
        status=patient_models.RelationshipStatus.CONFIRMED,
    )
    # homer
    marge_homer_relationship = patient_factories.Relationship.create(
        type=patient_models.RelationshipType.objects.guardian_caregiver(),
        patient=homer_patient,
        caregiver=marge_caregiver_profile,
        status=patient_models.RelationshipStatus.CONFIRMED,
    )
    homer_self_relationship = patient_factories.Relationship.create(
        type=patient_models.RelationshipType.objects.self_type(),
        patient=patient_factories.Patient.create(legacy_id=52, ramq='TEST01161973'),
        caregiver=homer_caregiver_profile,
        status=patient_models.RelationshipStatus.CONFIRMED,
    )
    homer_pending_self_relationship = patient_factories.Relationship.create(
        type=patient_models.RelationshipType.objects.self_type(),
        patient=homer_patient,
        caregiver=homer_caregiver_profile,
        status=patient_models.RelationshipStatus.PENDING,
    )
    # bart
    marge_bart_relationship = patient_factories.Relationship.create(
        type=patient_models.RelationshipType.objects.guardian_caregiver(),
        patient=bart_patient,
        caregiver=marge_caregiver_profile,
        status=patient_models.RelationshipStatus.CONFIRMED,
    )
    bart_self_relationship = patient_factories.Relationship.create(
        type=patient_models.RelationshipType.objects.self_type(),
        patient=patient_factories.Patient.create(legacy_id=53, ramq='TEST01161974'),
        caregiver=bart_caregiver_profile,
        status=patient_models.RelationshipStatus.CONFIRMED,
    )
    bart_expired_self_relationship = patient_factories.Relationship.create(
        type=patient_models.RelationshipType.objects.self_type(),
        patient=bart_patient,
        caregiver=bart_caregiver_profile,
        status=patient_models.RelationshipStatus.EXPIRED,
    )
    # lisa
    homer_lisa_relationship = patient_factories.Relationship.create(
        type=patient_models.RelationshipType.objects.guardian_caregiver(),
        patient=lisa_patient,
        caregiver=homer_caregiver_profile,
        status=patient_models.RelationshipStatus.CONFIRMED,
    )
    lisa_self_relationship = patient_factories.Relationship.create(
        type=patient_models.RelationshipType.objects.self_type(),
        patient=patient_factories.Patient.create(legacy_id=54, ramq='TEST01161975'),
        caregiver=lisa_caregiver_profile,
        status=patient_models.RelationshipStatus.CONFIRMED,
    )
    lisa_pending_self_relationship = patient_factories.Relationship.create(
        type=patient_models.RelationshipType.objects.self_type(),
        patient=lisa_patient,
        caregiver=lisa_caregiver_profile,
        status=patient_models.RelationshipStatus.PENDING,
    )

    return {
        'marge_relationship': marge_self_relationship,
        'homer_relationship': homer_self_relationship,
        'bart_relationship': bart_self_relationship,
        'lisa_relationship': lisa_self_relationship,
        'marge_homer_relationship': marge_homer_relationship,
        'homer_pending_relationship': homer_pending_self_relationship,
        'marge_bart_relationship': marge_bart_relationship,
        'bart_expired_relationship': bart_expired_self_relationship,
        'homer_lisa_relationship': homer_lisa_relationship,
        'lisa_pending_relationship': lisa_pending_self_relationship,
    }


def _create_device_identifier_records() -> None:
    """Create legacy device identifier records."""
    legacy_factories.LegacyPatientDeviceIdentifierFactory.create()
    legacy_factories.LegacyPatientDeviceIdentifierFactory.create()
    legacy_factories.LegacyPatientDeviceIdentifierFactory.create(device_type=1)
    legacy_factories.LegacyPatientDeviceIdentifierFactory.create(device_type=1)
    legacy_factories.LegacyPatientDeviceIdentifierFactory.create(device_type=3)
    legacy_factories.LegacyPatientDeviceIdentifierFactory.create(device_type=3)


def _create_received_medical_records() -> None:
    """Create received medical records."""
    current_date = timezone.now().date()

    stats_factories.DailyPatientDataReceived.create(
        patient=patient_factories.Patient.create(legacy_id=51, ramq='TEST01161972', created_at=timezone.now()),
        appointments_received=5,
        documents_received=5,
        educational_materials_received=5,
        labs_received=5,
        questionnaires_received=5,
        action_date=current_date,
        last_lab_received=timezone.now(),
    )
    stats_factories.DailyPatientDataReceived.create(
        patient=patient_factories.Patient.create(legacy_id=52, ramq='TEST01161973', created_at=timezone.now()),
        appointments_received=10,
        documents_received=10,
        educational_materials_received=10,
        labs_received=10,
        questionnaires_received=10,
        action_date=current_date,
        last_lab_received=timezone.now(),
    )
    stats_factories.DailyPatientDataReceived.create(
        patient=patient_factories.Patient.create(legacy_id=53, ramq='TEST01161974', created_at=timezone.now()),
        appointments_received=15,
        documents_received=15,
        educational_materials_received=15,
        labs_received=15,
        questionnaires_received=15,
        action_date=current_date,
        last_lab_received=timezone.now(),
    )
    stats_factories.DailyPatientDataReceived.create(
        patient=patient_factories.Patient.create(legacy_id=54, ramq='TEST01161975', created_at=timezone.now()),
        appointments_received=5,
        documents_received=5,
        educational_materials_received=5,
        labs_received=15,
        questionnaires_received=5,
        action_date=current_date - dt.timedelta(days=1),
        last_lab_received=timezone.now(),
    )
    stats_factories.DailyPatientDataReceived.create(
        patient=patient_factories.Patient.create(legacy_id=51, ramq='TEST01161972', created_at=timezone.now()),
        appointments_received=15,
        documents_received=15,
        educational_materials_received=15,
        labs_received=15,
        questionnaires_received=15,
        action_date=current_date - dt.timedelta(days=2),
        last_lab_received=timezone.now() - dt.timedelta(days=8),
    )


def _create_app_activity_records() -> None:
    """Create user app activity records."""
    current_date = timezone.now().date()
    stats_factories.DailyUserAppActivity.create(
        action_by_user=User.objects.filter(username='marge').first(),
        count_logins=3,
        count_feedback=4,
        count_update_security_answers=5,
        count_update_passwords=6,
        action_date=current_date - dt.timedelta(days=2),
        last_login=timezone.now() - dt.timedelta(days=2),
    )
    stats_factories.DailyUserAppActivity.create(
        action_by_user=User.objects.filter(username='homer').first(),
        count_logins=5,
        count_feedback=6,
        count_update_security_answers=7,
        count_update_passwords=8,
        action_date=current_date - dt.timedelta(days=2),
        last_login=timezone.now() - dt.timedelta(days=2),
    )
    stats_factories.DailyUserAppActivity.create(
        action_by_user=User.objects.filter(username='marge').first(),
        count_logins=10,
        count_feedback=11,
        count_update_security_answers=12,
        count_update_passwords=13,
        action_date=current_date,
        last_login=timezone.now(),
    )
    stats_factories.DailyUserAppActivity.create(
        action_by_user=User.objects.filter(username='homer').first(),
        count_logins=3,
        count_feedback=4,
        count_update_security_answers=5,
        count_update_passwords=6,
        action_date=current_date,
        last_login=timezone.now(),
    )
    stats_factories.DailyUserAppActivity.create(
        action_by_user=User.objects.filter(username='bart').first(),
        count_logins=5,
        count_feedback=6,
        count_update_security_answers=7,
        count_update_passwords=8,
        action_date=current_date,
        last_login=timezone.now() - dt.timedelta(days=2),
    )

    # For testing users latest login year summary
    stats_factories.DailyUserAppActivity.create(
        action_by_user=User.objects.filter(username='lisa').first(),
        last_login=dt.datetime(2024, 8, 20, 10, 10, 10).astimezone(),
        count_logins=3,
        action_date=dt.datetime(2024, 8, 20, tzinfo=timezone.get_current_timezone()),
    )
    stats_factories.DailyUserAppActivity.create(
        action_by_user=User.objects.filter(username='fred').first(),
        last_login=dt.datetime(2023, 8, 20, 10, 10, 10).astimezone(),
        count_logins=5,
        action_date=dt.datetime(2023, 8, 20, tzinfo=timezone.get_current_timezone()),
    )
    stats_factories.DailyUserAppActivity.create(
        action_by_user=User.objects.filter(username='mona').first(),
        last_login=dt.datetime(2022, 8, 20, 10, 10, 10).astimezone(),
        count_logins=5,
        action_date=dt.datetime(2022, 8, 20, tzinfo=timezone.get_current_timezone()),
    )


def _create_user_patient_activity_records() -> None:
    """Create user-patient activity records."""
    current_date = timezone.now().date()

    homer_relationship = patient_models.Relationship.objects.filter(
        caregiver__user__username='homer',
    )[0]
    marge_relationship = patient_models.Relationship.objects.filter(
        caregiver__user__username='homer',
    )[0]
    bart_relationship = patient_models.Relationship.objects.filter(
        caregiver__user__username='bart',
    )[0]
    lisa_relationship = patient_models.Relationship.objects.filter(
        caregiver__user__username='lisa',
    )[0]

    stats_factories.DailyUserPatientActivity.create(
        user_relationship_to_patient=homer_relationship,
        action_by_user=homer_relationship.caregiver.user,
        patient=homer_relationship.patient,
        count_checkins=3,
        count_documents=4,
        count_educational_materials=5,
        count_questionnaires_complete=6,
        count_labs=7,
        action_date=current_date - dt.timedelta(days=2),
    )
    stats_factories.DailyUserPatientActivity.create(
        user_relationship_to_patient=marge_relationship,
        action_by_user=marge_relationship.caregiver.user,
        patient=marge_relationship.patient,
        count_checkins=10,
        count_documents=11,
        count_educational_materials=12,
        count_questionnaires_complete=13,
        count_labs=14,
        action_date=current_date - dt.timedelta(days=2),
    )
    stats_factories.DailyUserPatientActivity.create(
        user_relationship_to_patient=bart_relationship,
        action_by_user=bart_relationship.caregiver.user,
        patient=bart_relationship.patient,
        count_checkins=5,
        count_documents=6,
        count_educational_materials=7,
        count_questionnaires_complete=8,
        count_labs=9,
        action_date=current_date - dt.timedelta(days=1),
    )
    stats_factories.DailyUserPatientActivity.create(
        user_relationship_to_patient=lisa_relationship,
        action_by_user=lisa_relationship.caregiver.user,
        patient=lisa_relationship.patient,
        count_checkins=7,
        count_documents=8,
        count_educational_materials=9,
        count_questionnaires_complete=10,
        count_labs=11,
        action_date=current_date - dt.timedelta(days=1),
    )
    stats_factories.DailyUserPatientActivity.create(
        user_relationship_to_patient=marge_relationship,
        action_by_user=marge_relationship.caregiver.user,
        patient=marge_relationship.patient,
        count_checkins=1,
        count_documents=2,
        count_educational_materials=3,
        count_questionnaires_complete=4,
        count_labs=5,
        action_date=current_date - dt.timedelta(days=1),
    )
    stats_factories.DailyUserPatientActivity.create(
        user_relationship_to_patient=homer_relationship,
        action_by_user=homer_relationship.caregiver.user,
        patient=homer_relationship.patient,
        count_checkins=3,
        count_documents=4,
        count_educational_materials=5,
        count_questionnaires_complete=6,
        count_labs=7,
        action_date=current_date - dt.timedelta(days=1),
    )
    stats_factories.DailyUserPatientActivity.create(
        user_relationship_to_patient=marge_relationship,
        action_by_user=marge_relationship.caregiver.user,
        patient=marge_relationship.patient,
        count_checkins=1,
        count_documents=2,
        count_educational_materials=3,
        count_questionnaires_complete=4,
        count_labs=5,
        action_date=current_date,
    )
    stats_factories.DailyUserPatientActivity.create(
        user_relationship_to_patient=lisa_relationship,
        action_by_user=lisa_relationship.caregiver.user,
        patient=lisa_relationship.patient,
        count_checkins=3,
        count_documents=4,
        count_educational_materials=5,
        count_questionnaires_complete=6,
        count_labs=7,
        action_date=current_date,
    )


def _create_demographic_diagnosis_records() -> None:
    """Create demographic diagnosis records."""
    # Marge
    marge_patient = patient_models.Patient.objects.filter(legacy_id=51, ramq='TEST01161972')[0]
    legacy_marge_patient = legacy_factories.LegacyPatientFactory.create(patientsernum=marge_patient.legacy_id)
    legacy_factories.LegacyPatientControlFactory.create(
        patient=legacy_factories.LegacyPatientFactory.create(patientsernum=marge_patient.legacy_id),
    )
    legacy_factories.LegacyPatientHospitalIdentifierFactory.create(
        patient=legacy_marge_patient,
        hospital='RVH',
        mrn=1234567,
    )

    # Homer
    homer_patient = patient_models.Patient.objects.filter(legacy_id=52, ramq='TEST01161973')[0]
    legacy_homer_patient = legacy_factories.LegacyPatientFactory.create(patientsernum=homer_patient.legacy_id)
    legacy_factories.LegacyPatientControlFactory.create(
        patient=legacy_factories.LegacyPatientFactory.create(patientsernum=homer_patient.legacy_id),
    )
    legacy_factories.LegacyPatientHospitalIdentifierFactory.create(
        patient=legacy_homer_patient,
        hospital='MGH',
        mrn=1234568,
    )
    # Bart
    bart_patient = patient_models.Patient.objects.filter(legacy_id=53, ramq='TEST01161974')[0]
    legacy_bart_patient = legacy_factories.LegacyPatientFactory.create(patientsernum=bart_patient.legacy_id)
    legacy_factories.LegacyPatientControlFactory.create(
        patient=legacy_factories.LegacyPatientFactory.create(patientsernum=bart_patient.legacy_id),
    )
    legacy_factories.LegacyPatientHospitalIdentifierFactory.create(
        patient=legacy_bart_patient,
        hospital='SJH',
        mrn=1234569,
    )
    # Lisa
    lisa_patient = patient_models.Patient.objects.filter(legacy_id=54, ramq='TEST01161975')[0]
    legacy_lisa_patient = legacy_factories.LegacyPatientFactory.create(patientsernum=lisa_patient.legacy_id)
    legacy_factories.LegacyPatientControlFactory.create(
        patient=legacy_factories.LegacyPatientFactory.create(patientsernum=lisa_patient.legacy_id),
    )
    legacy_factories.LegacyPatientHospitalIdentifierFactory.create(
        patient=legacy_lisa_patient,
        hospital='OMI',
        mrn=1234570,
    )

    # Diagnosis
    legacy_factories.LegacyDiagnosisFactory.create(
        patient_ser_num=legacy_marge_patient,
        description_en='Test Diagnosis1',
        creation_date=timezone.now() - dt.timedelta(days=1),
    )
    legacy_factories.LegacyDiagnosisFactory.create(
        patient_ser_num=legacy_marge_patient,
        description_en='Test Diagnosis2',
        creation_date=timezone.now(),
    )
    legacy_factories.LegacyDiagnosisFactory.create(
        patient_ser_num=legacy_homer_patient,
        description_en='Test Diagnosis1',
        creation_date=timezone.now(),
    )
    legacy_factories.LegacyDiagnosisFactory.create(
        patient_ser_num=legacy_bart_patient,
        description_en='Test Diagnosis1',
        creation_date=timezone.now(),
    )
