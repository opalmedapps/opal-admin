import io
import zipfile

from openpyxl import load_workbook

from .. import utils

# Sample file contents
file_contents = {
    'file1.txt': b'Hello, this is file 1.',
    'file2.txt': b'This is the content of file 2.',
    'image.png': b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00',
}


def test_random_uuid_length() -> None:
    """Test the length of random uuid."""
    length = 30
    uuid = utils.generate_random_uuid(length)
    assert len(uuid) == length


def test_random_uuid_is_string() -> None:
    """Ensure the random uuid is string."""
    uuid = utils.generate_random_uuid(30)
    assert isinstance(uuid, str)


def test_random_uuid_is_unique() -> None:
    """Ensure the random uuid is unique."""
    uuid1 = utils.generate_random_uuid(30)
    uuid2 = utils.generate_random_uuid(30)
    assert uuid1 != uuid2


def test_random_registration_code_format() -> None:
    """Ensure if random registration code is alpha-numeric string and equal to the given length."""
    length = 10
    test_institution_code = 'XX'
    code = utils.generate_random_registration_code(test_institution_code, length)
    assert len(code) == length + len(test_institution_code)
    assert isinstance(code, str)
    assert code.isalnum()


def test_random_registration_code_is_unique() -> None:
    """Ensure the random registration code is unique."""
    code1 = utils.generate_random_registration_code('XX', 10)
    code2 = utils.generate_random_registration_code('XX', 10)
    assert code1 != code2


def test_qr_code() -> None:
    """Ensure a QR code can be built as an in-memory SVG image."""
    code = utils.qr_code('https://opalmedapps.ca')

    assert code.startswith(b'<?xml version=\'1.0\' encoding=\'UTF-8\'?>\n<svg xmlns:svg="http://www.w3.org/2000/svg"')


def test_create_zip_empty_list() -> None:
    """Ensure create_zip does not fail when it receives an empty list."""
    files: dict[str, bytes] = {}

    # Invoke the function with the empty list
    zip_bytes = utils.create_zip(files)

    # Ensure the returned object is a bytes instance
    assert isinstance(zip_bytes, bytes), 'The returned object should be a bytes object even the file list is empty.'

    # Ensure the zip buffer is not None or empty
    assert zip_bytes, 'The ZIP buffer should not be None or empty.'

    # Open the zip file from the bytes object
    with zipfile.ZipFile(io.BytesIO(zip_bytes), 'r') as zip_file:
        # Get the list of files in the zip archive
        zip_contents = zip_file.namelist()

        # Assert that the zip archive is empty
        assert not zip_contents, 'The zip archive should be empty when the file list is empty.'


def test_create_zip_success() -> None:
    """Ensure a ZIP file can be successfully created as an in-memory bytes object."""
    zip_bytes = utils.create_zip(file_contents)

    assert isinstance(zip_bytes, bytes), 'The returned object should be a bytes instance.'
    assert zip_bytes, 'The ZIP buffer should not be empty.'


def test_create_zip_contains_all_files() -> None:
    """Ensure that all expected files are in the zip archive."""
    zip_bytes = utils.create_zip(file_contents)

    # Open the zip file from the bytes object
    with zipfile.ZipFile(io.BytesIO(zip_bytes), 'r') as zip_file:
        # Get the list of files in the zip archive
        zip_contents = zip_file.namelist()

        expected_files = list(file_contents.keys())

        assert set(zip_contents) == set(expected_files), (
            f'ZIP archive contents {zip_contents} do not match expected {expected_files}.'
        )


def test_create_zip_contains_files_contents() -> None:
    """Ensure that in the ZIP archive the contents of each file are the same as the original ones."""
    zip_bytes = utils.create_zip(file_contents)

    # Open the zip file from the bytes object
    with zipfile.ZipFile(io.BytesIO(zip_bytes), 'r') as zip_file:
        expected_files = list(file_contents.keys())

        # Verify the contents of each file
        for filename in expected_files:
            with zip_file.open(filename) as zipped_file:
                content = zipped_file.read()
                assert content == file_contents[filename], f'Contents of {filename} do not match.'


def test_dict_to_xlsx_empty_workbook() -> None:
    """Ensure that an empty data input results in an empty workbook."""
    data: utils.WorkbookData = {}
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    assert len(workbook.sheetnames) == 1


def test_dict_to_xlsx_single_sheet_single_row() -> None:
    """Test a workbook with one sheet and one row of data."""
    data: utils.WorkbookData = {
        'Sheet1': [
            {'Name': 'Alice', 'Age': 30, 'City': 'New York'},
        ],
    }
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    assert workbook.sheetnames == ['Sheet1']
    sheet = workbook['Sheet1']
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert headers == ['Name', 'Age', 'City']
    row = [cell.value for cell in next(sheet.iter_rows(min_row=2))]
    assert row == ['Alice', 30, 'New York']


def test_dict_to_xlsx_multiple_sheets_multiple_rows() -> None:
    """Test a workbook with multiple sheets and multiple rows."""
    data: utils.WorkbookData = {
        'Employees': [
            {'Name': 'Alice', 'Age': 30, 'Department': 'HR'},
            {'Name': 'Bob', 'Age': 35, 'Department': 'Engineering'},
        ],
        'Products': [
            {'Product ID': 'P001', 'Product Name': 'Widget', 'Price': 19.99},
            {'Product ID': 'P002', 'Product Name': 'Gadget', 'Price': 29.99},
        ],
    }
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    assert set(workbook.sheetnames) == {'Employees', 'Products'}

    # Test Employees sheet
    employees_sheet = workbook['Employees']
    headers = [cell.value for cell in next(employees_sheet.iter_rows(max_row=1))]
    assert headers == ['Name', 'Age', 'Department']
    rows = list(employees_sheet.iter_rows(min_row=2, values_only=True))
    assert rows == [
        ('Alice', 30, 'HR'),
        ('Bob', 35, 'Engineering'),
    ]

    # Test Products sheet
    products_sheet = workbook['Products']
    headers = [cell.value for cell in next(products_sheet.iter_rows(max_row=1))]
    assert headers == ['Product ID', 'Product Name', 'Price']
    rows = list(products_sheet.iter_rows(min_row=2, values_only=True))
    assert rows == [
        ('P001', 'Widget', 19.99),
        ('P002', 'Gadget', 29.99),
    ]


def test_dict_to_xlsx_missing_keys_in_rows() -> None:
    """Test handling of rows with missing keys."""
    data: utils.WorkbookData = {
        'Sheet1': [
            {'Name': 'Alice', 'Age': 30},
            {'Name': 'Bob'},
        ],
    }
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook['Sheet1']
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert headers == ['Name', 'Age']
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert rows == [
        ('Alice', 30),
        ('Bob', None),  # Missing value should be a None
    ]


def test_dict_to_xlsx_non_string_values() -> None:
    """Test handling of non-string values like integers and floats."""
    data: utils.WorkbookData = {
        'Numbers': [
            {'Integer': 42, 'Float': 1.05, 'Boolean': True},
        ],
    }
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook['Numbers']
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert headers == ['Integer', 'Float', 'Boolean']
    row = [cell.value for cell in next(sheet.iter_rows(min_row=2))]
    assert row == [42, 1.05, True]


def test_dict_to_xlsx_empty_sheet() -> None:
    """Test a workbook with a sheet that has no data."""
    data: utils.WorkbookData = {
        'EmptySheet': [],
    }
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    assert 'EmptySheet' in workbook.sheetnames
    sheet = workbook['EmptySheet']
    rows = list(sheet.iter_rows(values_only=True))
    assert not rows


def test_dict_to_xlsx_special_characters_in_sheet_name() -> None:
    """Ensure that special characters in sheet names are handled."""
    data: utils.WorkbookData = {
        'Data & Analysis': [
            {'Metric': 'Engagement', 'Value': 75.5},
        ],
    }
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    assert 'Data & Analysis' in workbook.sheetnames


def test_dict_to_xlsx_long_sheet_name() -> None:
    """Ensure that long sheet names are truncated appropriately."""
    long_sheet_name = 'A' * 40  # Exceeds Excel's 31-character limit
    data: utils.WorkbookData = {
        long_sheet_name: [
            {'Data': 'Test'},
        ],
    }
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    truncated_sheet_name = long_sheet_name[:utils.SHEET_TITLE_MAX_LENGTH]
    assert truncated_sheet_name in workbook.sheetnames


def test_dict_to_xlsx_invalid_sheet_name_characters_removed() -> None:
    """Ensure that invalid characters in sheet names are removed."""
    invalid_sheet_name = r'Invalid:/\\*?[]'
    data: utils.WorkbookData = {
        invalid_sheet_name: [
            {'Data': 'Test'},
        ],
    }
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    assert 'Invalid' in workbook.sheetnames


def test_dict_to_xlsx_empty_sheet_name() -> None:
    """Ensure that an empty sheet name is handled without errors."""
    invalid_sheet_name = r':/\\*?[]'
    data: utils.WorkbookData = {
        invalid_sheet_name: [
            {'Data': 'Test'},
        ],
    }
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    assert 'Sheet' in workbook.sheetnames


def test_dict_to_xlsx_none_values_in_data() -> None:
    """Ensure that None values in data are handled properly."""
    data: utils.WorkbookData = {
        'Sheet1': [
            {'Name': 'Alice', 'Age': None},
        ],
    }
    xlsx_bytes = utils.dict_to_xlsx(data)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook['Sheet1']
    row = [cell.value for cell in next(sheet.iter_rows(min_row=2))]
    assert row == ['Alice', None]
